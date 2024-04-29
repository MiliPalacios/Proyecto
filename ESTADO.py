import streamlit as st 
import numpy as np
import pandas as pd
from io import BytesIO
import funciones as f
from pages import pag1,pag2
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border
from openpyxl.drawing.image import Image as image
from PIL import Image
from fpdf import FPDF
import base64
st.sidebar.header("Paginas")
pages={
    "Extra1": pag1,
    "Extra2":pag2
}
selection=st.sidebar.radio("Ir a",list(pages.keys()))
page=pages[selection]
st.title("RESUMEN ESTADO DE CUENTA")
st.write("Esta app fue hecha con el fin de facilitar los procedimientos internos en la tesoreria de un conjunto de residencias")
st.warning("APP AUN DESARROLLANDOSE: no exite contenido en las paginas anexas")
# Subir base desde el computador
#try:
#    archivo_base = st.file_uploader('Subir base de datos',type=['xlsx'])
#except:
#    st.warning("Se debe subir un archivo de EXCEL",icon="🔴")
    
#if archivo_base is not None:
#    st.success("Archivo subido correctamente",icon="✅")
    # Lectura de la base
#    datos = pd.read_excel(archivo_base)    
    #Transformar datos de fecha a datos de tiempo
datos=pd.read_excel("datasets\Movimiento.xlsx")
for i in range (len(datos.index)) :
    datos["Fecha"][i]=pd.to_datetime(f.texto_a_fechas(datos["Fecha"][i])) 
datos=datos.fillna(int("0"))
st.subheader("La base de datos es:")
st.write(datos)
    #AGREGAR UN MENU DE OPCIONES
opciones=['Analisis general',"Analisis por lote",'Analisis por dia','Filtrar información']
opcion = st.selectbox('¿Qué deseas hacer?',opciones)
if opcion=='Filtrar información':
        #FILTRAR INFORMACION POR CLASES EN COLUMNAS
        st.subheader('Filtrar información')
        columnas=list(datos.columns)
        columna = st.selectbox('Selecciona una opción',columnas)
        valores=datos[columna].unique()
        valor = st.selectbox('Selecciona una opción',valores)
        datos_f=datos[datos[columna]==valor]
        st.write(datos_f)
if opcion=='Analisis general':
        #Data frame depurado
        st.subheader('INGRESOS VS FACTURAS')
        st.write('A continuación se presenta una gráfica de los montos facturados y no facturados. Objetivo: dar un vistazo general al movimiento de ADMINISTRACIÓN.')
        #Eliminar valores negativos
        datos_importantes=datos[datos["Tipo de transacción"].isin(["Crédito"])]

        #Eliminar columnas
        datos_importantes=datos_importantes.drop(columns=["Tipo de transacción","Oficina","Concepto","Documento","Saldo","Hora","ORDENANTE","CUENTA ORIGEN","DESCRIPCION BANCO","DETALLE PAGO","BANCO","OBSERVACION","OBSERVACIONES","ESTADO CONTABILIDAD","N. de comprobante"])
       
        #GRAFICO FACTURAS VS INGRESOS
        dias=datos_importantes["Fecha"].unique()
        x = np.arange(len(dias))
        ancho=0.25
        datos_importantes=datos[datos["Tipo de transacción"].isin(["Crédito"])]
        dias=datos_importantes["Fecha"].unique()
        ingresos=[]
        facturado=[]
        for i in range(len(dias)):
            dia=dias[i]
            datos_i=datos_importantes[datos_importantes["Fecha"].isin([dia])].drop(columns=["LOTE","MES","ESTADO CONTABILIDAD","CUENTA ORIGEN","DESCRIPCION BANCO","OBSERVACIONES","OBSERVACION","Tipo de transacción","Oficina","Concepto","Documento","Saldo","DETALLE PAGO","BANCO","N. de comprobante"])
            valores_ingresos_i=datos_i["Monto"].sum()
            valores_facturados_i=datos_i[datos_i["FACTURA"]!=0]["Monto"].sum()
            ingresos.append(valores_ingresos_i)
            facturado.append(valores_facturados_i)

        df=pd.DataFrame({
            "Ingresos" : ingresos,
            "Facturado": facturado,
            })
        st.line_chart(df,color=["#F50404","#002EB1"])
        col1,col2=st.columns(2)
        col1.write("En caso de existir inconsistencias, revisar el dataframe a continuación:")
        col1.write(datos_importantes)

        col2.subheader('INGRESOS VS EGRESOS')
        datos_i_e=datos.loc[:,["Fecha","Tipo de transacción","Monto"]]
        tabla_i_e=datos_i_e.pivot_table(index="Tipo de transacción", values="Monto", aggfunc= lambda x:sum(x))
        col2.write(tabla_i_e)
if opcion=='Analisis por dia':
        #Ingresos y egresos por día     
        fechas=datos["Fecha"].unique()
        datos_i_e=datos.loc[:,["Fecha","Tipo de transacción","Monto"]]
        fecha=st.selectbox('Seleccione una opción: ', fechas)
        col1,col2=st.columns(2)
       #Ingresos y egresos por día     
        col1.subheader('TRANSACCIONES')
        col1.info("Seccion en CORRECCION: dias 0")
        col1.info("Para ejemplo rápido seleccionar: 2024-01-03")
        datos_d=datos_i_e[datos_i_e["Fecha"].isin([fecha])]
        tabla_d=datos_d.pivot_table(index="Tipo de transacción",columns="Fecha", values="Monto", aggfunc= lambda x:sum(x))
        col1.write("Totales repecto al tipo de transacción")
        col1.write(tabla_d)
        col1.write("Aquí se presenta el valor de los montos depositados. Objetivo: identificar valores no usuales en las transacciones")
        col1.line_chart(datos_d["Monto"],color="#08FF01")    
        
        col2.subheader('INGRESOS VS FACTURAS')
        col2.write("Aqui usted puede verificar los ingresos que aun no han sido facturados")
        datos_dia=datos[datos["Fecha"].isin([fecha])].drop(columns=["LOTE","MES","ESTADO CONTABILIDAD","CUENTA ORIGEN","DESCRIPCION BANCO","OBSERVACIONES","OBSERVACION","Oficina","Concepto","Documento","Saldo","DETALLE PAGO","BANCO","N. de comprobante"])
        datos_dia=datos_dia[datos_dia["Tipo de transacción"].isin(["Crédito"])]
        col2.write(datos_dia)
        valores_ingresos_d=datos_dia["Monto"].sum()
        valores_facturados_d=datos_dia[datos_dia["FACTURA"]!=0]["Monto"].sum()
        y2=[valores_ingresos_d,valores_facturados_d]
        x2=[1,2]
        plt.figure(figsize=(10, 6))
        plt.title(f"CONTROL DEL REGISTRO DE INGRESOS DIA {fecha}")
        barras2=plt.bar(x2,y2,width=0.8,edgecolor="k",color=["g","b"])
        plt.ylabel('USD')
        plt.xticks([1,2],["INGRESO","MONTO FACTURADO"])
        for barra in barras2:
            altura = barra.get_height()
            plt.text(barra.get_x()+barra.get_width()/2, altura/2, altura, ha = 'center',va="bottom")
        plt.savefig('images/i_vs_f_dia.png')
        image=Image.open('./images/i_vs_f_dia.png')
        col2.image(image,caption="Objetivo: dar un vistazo más cercano a las actualizaciones de ADMINISTRACIÓN.")
if opcion=="Analisis por lote":
        datos["LOTE"]=datos["LOTE"].apply(f.lotes_sin_nombre)
        datos=datos.sort_values(by=["LOTE"])
        lotes=list(datos["LOTE"].unique())
        lote=st.selectbox("Escoja el lote:", lotes)
        datos_lote=datos[datos["Tipo de transacción"]!= "Débito"]
        datos_lote=datos_lote[datos_lote["LOTE"].isin([lote])].drop(columns=["LOTE","MES","ESTADO CONTABILIDAD","CUENTA ORIGEN","DESCRIPCION BANCO","OBSERVACIONES","OBSERVACION","Tipo de transacción","Oficina","Concepto","Documento","Saldo","DETALLE PAGO","BANCO","N. de comprobante"])
        #Graficos
        valores_ingresos_l=datos_lote["Monto"].sum()
        facturas_l=datos_lote[datos_lote["FACTURA"]!=0]
        valores_facturados_l=facturas_l["Monto"].sum()
        y1=[valores_ingresos_l,valores_facturados_l]
        x1=[1,2]
        plt.figure(figsize=(10, 6))
        plt.title(f"CONTROL DEL REGISTRO DE INGRESOS LOTE {lote}")
        barras1=plt.bar(x1,y1,width=0.8,edgecolor="k",color=["g","b"])
        plt.ylabel('Cantidad')
        plt.xticks([1,2],["INGRESOS","FACTURAS"])
        for barra in barras1:
            altura = barra.get_height()
            plt.text(barra.get_x()+barra.get_width()/2, altura/2, altura, ha = 'center',va="bottom")
        plt.savefig(f'images/i_vs_f_lote.png')
        plt.show()
        image=Image.open('./images/i_vs_f_lote.png')
        st.image(image,caption="Objetivo: dar un vistazo más cercano a las actualizaciones de ADMINISTRACIÓN.")

        #Análisis EXCEL
        wb=Workbook()
        ws=wb.active
        ws["E1"]=f"REPORTE DEL LOTE {lote}"
        ws['E1'].font = Font(name='Amercian Typewriter',size=20,bold=True,italic=True,color='139911')
        ws.row_dimensions[1].height = 30
        ws['E1'].alignment = Alignment(horizontal='center',vertical='center')    
        list=["B","C","D","E","F","G","H"]
        #Encabezado tabla
        for j in range (len(list)):
            ws.column_dimensions[f'{list[j]}'].width=20
            ws[f"{list[j]}4"]=datos_lote.columns[j]
            ws[f"{list[j]}4"].font = Font(name='Times',size=12,bold=True,italic=False,color='243783')
            ws[f"{list[j]}4"].alignment = Alignment(horizontal='center',vertical='center')
        #Datos Tabla
        aux=0
        for i in range (5,5+len(datos_lote.index)):
            ws[f"B{i}"]=str(pd.to_datetime(datos_lote["Fecha"].values[aux]))
            ws[f"B{i}"].alignment = Alignment(horizontal='center',vertical='center')
            ws[f"C{i}"]=datos_lote["Hora"].values[aux]
            ws[f"C{i}"].alignment = Alignment(horizontal='center',vertical='center')
            ws[f"D{i}"]=datos_lote["Monto"].values[aux]
            ws[f"D{i}"].alignment = Alignment(horizontal='center',vertical='center')
            ws[f"E{i}"]=datos_lote["ORDENANTE"].values[aux]
            ws[f"E{i}"].alignment = Alignment(horizontal='center',vertical='center')
            ws[f"F{i}"]=datos_lote["FACTURA"].values[aux]
            ws[f"F{i}"].alignment = Alignment(horizontal='center',vertical='center')
            ws[f"G{i}"]=str(pd.to_datetime(datos_lote["FECHA FACTURA"].values[aux]))
            ws[f"G{i}"].alignment = Alignment(horizontal='center',vertical='center')
            ws[f"H{i}"]=datos_lote["MOTIVO"].values[aux]
            ws[f"H{i}"].alignment = Alignment(horizontal='center',vertical='center')
            aux+=1
        #Agregar imagenes
        #img = image('images/i_vs_f_lote.png')
        #ws.add_image(img,"B10")
        #DESCARGAR EL EXCEL
        excel_file=f.descargar_excel(wb)
        st.download_button(label="Descargar Excel",data=excel_file,file_name=f"Reporte_lote_{lote}.xlsx")
        
        #Reporte en PDF
        #Funciones principales
        class PDFWithBackground(FPDF):
                    def __init__(self):
                        super().__init__()
                        self.background = None

                    def set_background(self, image_path):
                        self.background = image_path

                    def add_page(self, orientation=''):
                        super().add_page(orientation)
                        if self.background:
                            self.image(self.background, 0, 0, self.w, self.h)

                    def footer(self):
                        # Posición a 1.5 cm desde el fondo
                        self.set_y(-15)
                        # Configurar la fuente para el pie de página
                        self.set_font('Arial', 'I', 8)
                        # Número de página
                        self.cell(0, 10, 'Página ' + str(self.page_no()), 0, 0, 'C')
        #Crear el PDF
        pdf=PDFWithBackground()
        #Encabezado pagina
        pdf.set_background("images/background.jpeg")
        pdf.add_page()
        pdf.set_y(10)
        pdf.set_font("Times",size=10)#Arial, Times, Courier
        pdf.cell(0,0,"RESUMEN AUTOMÁTICO",0,1,"L")

        pdf.set_y(10)
        pdf.set_x(90)
        pdf.set_font("Times",style="I",size=10)#Arial, Times, Courier
        pdf.cell(0,0,"Elaborado por: Milena Palacios",0,1)

        pdf.set_y(22)
        pdf.set_x(15)
        pdf.set_font("Times",style="B",size=13)#Arial, Times, Courier
        pdf.cell(0,0,"URBANIZACIÓN DEL EJÉRCITO NACIONAL",0,1)

        pdf.set_y(40)
        pdf.set_x(20)
        pdf.set_font("Times",style="B",size=55)#Arial, Times, Courier
        pdf.cell(0,0,f"LOTE {lote}",0,1)

        pdf.set_y(55)
        pdf.set_x(25)
        pdf.set_font("Times",style="B",size=11)#Arial, Times, Courier
        pdf.cell(0,0,"INFORMACIÓN GENERAL (en desarrollo)",0,1)

        pdf.set_y(65)
        pdf.set_x(10)
        pdf.set_font("Times",size=12)#Arial, Times, Courier
        pdf.cell(0,0,"Ubicación:",0,1)

        pdf.set_y(70)
        pdf.set_x(10)
        pdf.set_font("Times",size=12)#Arial, Times, Courier
        pdf.cell(0,0,"Residente:",0,1)

        pdf.set_y(75)
        pdf.set_x(10)
        pdf.set_font("Times",size=12)#Arial, Times, Courier
        pdf.cell(0,0,"CI:",0,1)

        pdf.set_y(80)
        pdf.set_x(10)
        pdf.set_font("Times",size=12)#Arial, Times, Courier
        pdf.cell(0,0,"Correo:",0,1)

        pdf.set_y(85)
        pdf.set_x(10)
        pdf.set_font("Times",size=12)#Arial, Times, Courier
        pdf.cell(0,0,"Contacto:",0,1)

        #Grafico
        pdf.image(f'images/i_vs_f_lote.png',x=91,y=100,w=120,h=65)
        #Encabezado tabla
        pdf.set_y(170)
        pdf.set_x(20)
        pdf.set_font("Times",style="B",size=15)
        pdf.cell(w=35,h=10,txt="Fecha",border=1, align="C",fill=0)
        pdf.cell(w=20,h=10,txt="Hora",border=1, align="C",fill=0)
        pdf.cell(w=20,h=10,txt="Monto",border=1, align="C",fill=0)
        pdf.cell(w=75,h=10,txt="ORDENANTE",border=1, align="C",fill=0)
        pdf.cell(w=35,h=10,txt="FACTURA",border=1, align="C",fill=0)

        #DATOS TABLA
        new_posicion=180
        pdf.set_y(new_posicion)
        pdf.set_x(20)
        altura=8
        for i in range (len(datos_lote.index)):
            pdf.set_font("Times",size=10)
            pdf.cell(w=35,h=altura,txt=str(pd.to_datetime(datos_lote["Fecha"].values[i])),border=1, align="C",fill=0)
            pdf.cell(w=20,h=altura,txt=str(datos_lote["Hora"].values[i]),border=1, align="C",fill=0)
            pdf.cell(w=20,h=altura,txt=str(datos_lote["Monto"].values[i]),border=1, align="C",fill=0)
            pdf.cell(w=75,h=altura,txt=str(datos_lote["ORDENANTE"].values[i]),border=1, align="C",fill=0)
            pdf.cell(w=35,h=altura,txt=str(datos_lote["FACTURA"].values[i]),border=1, align="C",fill=0)
            new_posicion= new_posicion + altura
            #Poner información conjunta
            if new_posicion<=270:
                pdf.set_y(new_posicion)
                pdf.set_x(20)
            else:
                new_posicion=10
                pdf.add_page()
                pdf.set_y(new_posicion)
                pdf.set_x(20)
        pdf.set_y(275)
        pdf.set_font("Times",style="I",size=14)#Arial, Times, Courier
        pdf.cell(0,0,"¡Juntos trabajamos por el bienestar!",0,1,"C")
        #Descargar PDF
        def create_download_link(val, filename):
            b64 = base64.b64encode(val)
            return f'<a href="data:application/octet-stream;base64,{b64.decode()}" download="{filename}.pdf">Descargar reporte en PDF</a>'
        html=create_download_link(pdf.output(dest="S").encode("latin-1"), f"reportes/Reporte_lote_{lote}")
        st.divider()
        col_1,col_2 = st.columns(2)
        col_1.markdown(html, unsafe_allow_html=True)